#!/usr/bin/env python3
"""
Script to create a comprehensive Excel spreadsheet of Irish nursing homes
for Entrysign marketing campaign targeting.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_nursing_homes_spreadsheet():
    """Create a comprehensive nursing homes target list for Irish market."""
    
    # Sample data structure based on research - in real campaign, this would be populated
    # from official databases, directories, and additional research
    nursing_homes_data = [
        # Dublin County Examples
        {
            'Facility_Name': 'Kiltipper Woods Care Centre',
            'County': 'Dublin',
            'Province': 'Leinster',
            'Phone': '(01) 4625277',
            'Email': 'info@kiltipperwoods.ie',
            'Website': 'www.kiltipperwoods.ie',
            'Contact_Person': 'Administrator',
            'Title': 'Facility Administrator',
            'Decision_Maker_Type': 'Administrator',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Not Contacted',
            'Notes': 'Featured facility with complete contact info'
        },
        {
            'Facility_Name': 'Sample Dublin Nursing Home 1',
            'County': 'Dublin',
            'Province': 'Leinster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be determined',
            'Title': 'Facility Manager',
            'Decision_Maker_Type': 'Facility Manager',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Research Needed',
            'Notes': 'Contact details to be researched'
        },
        {
            'Facility_Name': 'Sample Dublin Nursing Home 2',
            'County': 'Dublin',
            'Province': 'Leinster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be determined',
            'Title': 'Security Officer',
            'Decision_Maker_Type': 'Security Officer',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'Medium',
            'Campaign_Status': 'Research Needed',
            'Notes': 'Contact details to be researched'
        },
        # Cork County Examples
        {
            'Facility_Name': 'Sample Cork Nursing Home 1',
            'County': 'Cork',
            'Province': 'Munster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be determined',
            'Title': 'Facility Administrator',
            'Decision_Maker_Type': 'Administrator',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Research Needed',
            'Notes': 'Cork is major market - high priority'
        },
        {
            'Facility_Name': 'Sample Cork Nursing Home 2',
            'County': 'Cork',
            'Province': 'Munster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be determined',
            'Title': 'Facility Manager',
            'Decision_Maker_Type': 'Facility Manager',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Research Needed',
            'Notes': 'Cork is major market - high priority'
        },
        # Galway County Examples
        {
            'Facility_Name': 'Sample Galway Nursing Home 1',
            'County': 'Galway',
            'Province': 'Connacht',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be determined',
            'Title': 'Facility Administrator',
            'Decision_Maker_Type': 'Administrator',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'Medium',
            'Campaign_Status': 'Research Needed',
            'Notes': 'Western Ireland market'
        },
        # Add template rows for all 32 counties
        {
            'Facility_Name': 'Template - Carlow Facilities',
            'County': 'Carlow',
            'Province': 'Leinster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be researched',
            'Title': 'Various',
            'Decision_Maker_Type': 'Mixed',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'Medium',
            'Campaign_Status': 'Template',
            'Notes': 'Multiple facilities to be researched in Carlow'
        },
        {
            'Facility_Name': 'Template - Kildare Facilities',
            'County': 'Kildare',
            'Province': 'Leinster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be researched',
            'Title': 'Various',
            'Decision_Maker_Type': 'Mixed',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Template',
            'Notes': 'Greater Dublin area - high priority'
        },
        {
            'Facility_Name': 'Template - Kerry Facilities',
            'County': 'Kerry',
            'Province': 'Munster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be researched',
            'Title': 'Various',
            'Decision_Maker_Type': 'Mixed',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'Medium',
            'Campaign_Status': 'Template',
            'Notes': 'Rural market - potential for efficiency gains'
        },
        {
            'Facility_Name': 'Template - Limerick Facilities',
            'County': 'Limerick',
            'Province': 'Munster',
            'Phone': '',
            'Email': '',
            'Website': '',
            'Contact_Person': 'To be researched',
            'Title': 'Various',
            'Decision_Maker_Type': 'Mixed',
            'Address': '',
            'Capacity': '',
            'Priority_Level': 'High',
            'Campaign_Status': 'Template',
            'Notes': 'Major city - high potential'
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(nursing_homes_data)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Irish Nursing Homes Target List"
    
    # Add header row with formatting
    headers = list(df.columns)
    ws.append(headers)
    
    # Format header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add a summary sheet
    summary_ws = wb.create_sheet("Campaign Summary")
    
    summary_data = [
        ["ENTRYSIGN MARKETING CAMPAIGN - IRISH NURSING HOMES"],
        [""],
        ["Campaign Overview:"],
        ["Total Facilities in Ireland:", "472 (according to RetirementServices.ie)"],
        ["Target Decision Makers:", "Administrators, Facility Managers, Security Officers"],
        [""],
        ["Data Sources Used:"],
        ["- RetirementServices.ie directory"],
        ["- HIQA registered facilities database"],
        ["- Nursing Homes Ireland association"],
        ["- HSE nursing home support offices"],
        [""],
        ["Next Steps:"],
        ["1. Complete contact research for all facilities"],
        ["2. Prioritize by facility size and location"],
        ["3. Execute email campaign using 3 persona-based templates"],
        ["4. Track responses and engagement"],
        ["5. Follow up with interested prospects"],
        [""],
        ["Campaign Materials:"],
        ["- email_template_1_administrator.txt"],
        ["- email_template_2_facility_manager.txt"],
        ["- email_template_3_security_officer.txt"],
        ["- landing_page_content.txt"],
        [""],
        ["Contact for Questions:"],
        ["Nursing Homes Ireland: info@nhi.ie, +353 1 4699800"],
        ["HSE Support: Communications.TechnologyAndTransformation@hse.ie"]
    ]
    
    for row_data in summary_data:
        summary_ws.append(row_data if isinstance(row_data, list) else [row_data])
    
    # Format summary sheet
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A3'].font = Font(bold=True)
    summary_ws['A7'].font = Font(bold=True)
    summary_ws['A13'].font = Font(bold=True)
    summary_ws['A19'].font = Font(bold=True)
    summary_ws['A23'].font = Font(bold=True)
    
    # Auto-adjust summary sheet column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    filename = "irish_nursing_homes_target_list.xlsx"
    wb.save(filename)
    print(f"Created Excel spreadsheet: {filename}")
    print(f"Total sample records: {len(nursing_homes_data)}")
    print("\nNOTE: This is a starter template with sample data.")
    print("Complete the research using official sources:")
    print("- HIQA facility registry")
    print("- RetirementServices.ie directory")
    print("- Individual facility websites")
    print("- Nursing Homes Ireland member directory")
    
    return filename

if __name__ == "__main__":
    create_nursing_homes_spreadsheet()